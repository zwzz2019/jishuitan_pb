#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
医院医生排班程序 v2
新增: MR编号、班次间隔、追加不能MR、半效辅助、一周最多1MR
"""

import pandas as pd
import re
from dataclasses import dataclass, field
from openpyxl import load_workbook

# ============================================================
# 常量
# ============================================================
DAYS = ['周一', '周二', '周三', '周四', '周五']
DAY_COLS = [3, 4, 5, 6, 7]

DAILY_TARGETS = {
    '周一': (9, 9, 6), '周二': (7, 7, 6),
    '周三': (7, 7, 6), '周四': (7, 7, 6), '周五': (6, 7, 6),
}

SENIORS = ['陈绪珠', '程晓光', '白荣杰', '闫东', '尤玉华', '赵涛', '翁磊', '孙晶']
ASSISTANTS = [
    '陈垒', '季爱华', '陈建安', '杨柳', '赵海清', '王苹',
    '师欣瑶', '王露瑶', '王可心', '李婉淑', '李柯莹', '王馨', '范闻宇',
    '曾德威', '郝应龙', '林晨', '陈娅', '王业学'
]
NO_MR_ASSIST = [
    '杨柳', '王苹', '师欣瑶', '王露瑶', '王可心',
    '李婉淑', '李柯莹', '王馨', '范闻宇'
]
# 半效辅助人：本周变更——原 7 人全部转为全效辅助，故清空。
# 注意：清空后 phase5/phase6 中针对 HALF_EFFICIENCY 的成对/独立兜底分支不再触发，
# 这是有意为之；恢复半效需要把名单填回。
HALF_EFFICIENCY = []
# 全效辅助人
FULL_EFFICIENCY = [a for a in ASSISTANTS if a not in HALF_EFFICIENCY]

NEVER_RECEIVE_ASSIST = ['刘彦含', '郑璇']
ASSIST_PRIORITY_HIGHEST = ['裴京哲']
ASSIST_PRIORITY_SECOND = ['程克斌', '娄露馨', '徐黎', '李新民', '李娜']

APPENDABLE_SHIFTS = ['胃肠', '开药+', '输卵管']

# 明确有多个MR配额的人(不受"一周最多1MR"限制)
MULTI_MR_ALLOWED = ['林晨', '徐黎', '李新民', '李娜', '郑璇']

# 有回急白的人(本周优先X,不排MR)
HUI_JI_BAI_PEOPLE = ['苏永彬', '杨帆', '胥晓明', '詹惠荔', '刘超', '曹祯', '陈建安']


# ============================================================
# 工具函数
# ============================================================
def is_empty(val):
    return pd.isna(val) or str(val).strip() in ('', 'nan')

def is_pure_jieru(val):
    if pd.isna(val): return False
    return str(val).strip() == '介入'

def is_appendable(val):
    if pd.isna(val): return False
    return str(val).strip() in APPENDABLE_SHIFTS

def get_target(day_idx, shift_type):
    x_t, ct_t, mr_t = DAILY_TARGETS[DAYS[day_idx]]
    return {'X': x_t, 'CT': ct_t, 'MR': mr_t}[shift_type]

def get_upper_limit(day_idx, shift_type):
    if shift_type == 'MR': return 6
    if day_idx == 0: return get_target(0, shift_type)  # 周一恰好
    # 周二~五: 没有硬上限，用软约束在评分中控制均衡
    return 99


# ============================================================
# 数据结构
# ============================================================
@dataclass
class Doctor:
    name: str
    abbr: str
    row_idx: int
    category: str
    quota: dict = field(default_factory=dict)
    schedule: list = field(default_factory=list)
    original: list = field(default_factory=list)
    needs_scheduling: bool = False
    constraints: dict = field(default_factory=dict)
    mr_count_this_week: int = 0  # 本周已分配的MR数量
    max_mr: int = 1  # 本周最多MR数量


# ============================================================
# 数据加载
# ============================================================
def parse_quota(col3_str, name):
    if pd.isna(col3_str) or str(col3_str).strip() == '':
        return {}, False, {}, 1
    s = str(col3_str).strip()

    # 特殊硬编码
    if name in ['陈绪珠', '闫东', '尤玉华', '马毅民']:
        return {}, False, {}, 1
    if name in ['曾德威', '郝应龙', '陈娅', '王业学']:
        return {'X': 1, 'CT': 2, 'MR': 1}, True, {}, 1
    if name == '林晨':
        return {'X': 1, 'MR': 3}, True, {}, 3
    if name == '赵涛':
        return {'X': 2}, True, {}, 0
    if name == '王玲':
        return {'X': 1, 'MR': 1}, True, {'fixed_days': [3, 4]}, 1

    if not re.search(r'\d', s):
        return {}, False, {}, 1

    constraints = {}
    day_map = {'一': 0, '二': 1, '三': 2, '四': 3, '五': 4}

    # rules.md §四：日期指定可写 '在周X' 或 '在星期X' 两种形式
    pin_match = re.search(r'(\d*)个?在(?:周|星期)([一二三四五])', s)
    if pin_match:
        constraints['pin_day'] = day_map[pin_match.group(2)]

    s_clean = re.sub(r'[（(].*?[)）]', '', s)
    s_clean = re.sub(r'[，,]?\s*\d*个?在(?:周|星期)[一二三四五]', '', s_clean).strip('，, ')

    matches = re.findall(r'(\d+)\s*(X|CT|MR)', s_clean, re.IGNORECASE)
    if matches:
        quota = {}
        max_mr = 1
        for num, st in matches:
            quota[st.upper()] = int(num)
        mr_val = quota.get('MR', 0)
        if mr_val > 1:
            max_mr = mr_val
        elif mr_val == 1:
            max_mr = 1
        else:
            max_mr = 0 if 'MR' not in quota else 1
        return quota, True, constraints, max_mr

    num_match = re.search(r'(\d+)', s_clean)
    if num_match:
        return {'total': int(num_match.group(1))}, True, constraints, 1

    first_num = re.search(r'(\d+)', s)
    if first_num:
        return {'total': int(first_num.group(1))}, True, constraints, 1

    return {}, False, {}, 1


def load_all(filepath):
    df = pd.read_excel(filepath, header=None)
    doctors = {}

    for i in range(2, len(df)):
        name = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ''
        if not name or name == 'nan' or name == '马毅民':
            continue

        abbr = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) else ''
        col3 = df.iloc[i, 2]

        if name in SENIORS:
            category = 'senior'
        elif name in ASSISTANTS:
            category = 'assistant'
        else:
            category = 'normal'

        original = []
        for day_idx in range(5):
            val = df.iloc[i, DAY_COLS[day_idx]]
            original.append(val if pd.notna(val) else None)

        quota, needs_scheduling, constraints, max_mr = parse_quota(col3, name)

        # MULTI_MR_ALLOWED的人不受限制
        if name in MULTI_MR_ALLOWED:
            max_mr = quota.get('MR', 99) if 'MR' in quota else 99

        # 有MR盯机的人: max_mr=0 (盯机已占用MR名额)
        has_mr_dinji = False
        for day_idx in range(5):
            val = str(df.iloc[i, DAY_COLS[day_idx]]).strip() if pd.notna(df.iloc[i, DAY_COLS[day_idx]]) else ''
            if 'MR' in val and '盯机' in val:
                has_mr_dinji = True
                break
        if has_mr_dinji and name not in MULTI_MR_ALLOWED:
            max_mr = 0

        # 回急白的人: 不排MR
        has_hui_ji_bai = False
        for day_idx in range(7):  # 检查全周
            col_idx = DAY_COLS[0] + day_idx if day_idx < 5 else (8 if day_idx == 5 else 9)
            if col_idx < len(df.columns):
                val = str(df.iloc[i, col_idx]).strip() if pd.notna(df.iloc[i, col_idx]) else ''
                if val == '回急白':
                    has_hui_ji_bai = True
                    break
        if has_hui_ji_bai:
            max_mr = 0
            constraints['hui_ji_bai'] = True  # 标记优先排X

        # 个人 MR 禁排日（rules.md §九）
        # 李新民: 周一不排MR
        if name == '李新民':
            constraints['mr_blocked_days'] = [0]

        doctors[name] = Doctor(
            name=name, abbr=abbr, row_idx=i, category=category,
            quota=quota, schedule=[None]*5, original=original,
            needs_scheduling=needs_scheduling, constraints=constraints,
            max_mr=max_mr,
        )

    return df, doctors


def count_existing(df):
    counts = {d: {'X': 0, 'CT': 0, 'MR': 0} for d in range(5)}
    for i in range(2, len(df)):
        name = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ''
        if name == '马毅民' or not name:
            continue
        is_asst = name in ASSISTANTS
        for day_idx in range(5):
            val = str(df.iloc[i, DAY_COLS[day_idx]]).strip() if pd.notna(df.iloc[i, DAY_COLS[day_idx]]) else ''
            if val == 'X':
                counts[day_idx]['X'] += 1
            elif val == 'CT':
                counts[day_idx]['CT'] += 1
            elif val == 'MR':
                counts[day_idx]['MR'] += 1
            elif 'MR' in val and '盯机' in val and '辅助' not in val:
                # 纯辅助人的MR盯机不计入headcount(需配对)
                if not is_asst:
                    counts[day_idx]['MR'] += 1
    return counts


# ============================================================
# 分配引擎(含间隔约束和MR限制)
# ============================================================
def get_shift_on_day(doctor, day_idx):
    """获取某人某天的班种(含original)"""
    if doctor.schedule[day_idx] in ('X', 'CT', 'MR'):
        return doctor.schedule[day_idx]
    orig = doctor.original[day_idx]
    if orig and str(orig).strip() in ('X', 'CT', 'MR'):
        return str(orig).strip()
    return None


def check_consecutive(doctor, day_idx, shift_type):
    """检查是否会造成连续同班种(软性,返回penalty)"""
    penalty = 0
    # 前一天
    if day_idx > 0:
        prev = get_shift_on_day(doctor, day_idx - 1)
        if prev == shift_type:
            penalty += 20  # 强惩罚
    # 后一天
    if day_idx < 4:
        nxt = get_shift_on_day(doctor, day_idx + 1)
        if nxt == shift_type:
            penalty += 20  # 强惩罚
    return penalty


def find_best_day(shift_type, available_days, used_days, counts, doctor=None, extra_filter=None):
    """找缺口最大且不超限的天，考虑间隔和均衡"""
    candidates = []
    for d in available_days:
        if d in used_days:
            continue
        if counts[d][shift_type] >= get_upper_limit(d, shift_type):
            continue
        # 个人 MR 禁排日（rules.md §九，例如李新民周一不排MR）
        if doctor and shift_type == 'MR' and d in doctor.constraints.get('mr_blocked_days', ()):
            continue
        if extra_filter and not extra_filter(d):
            continue
        gap = get_target(d, shift_type) - counts[d][shift_type]
        penalty = 0
        if doctor:
            penalty = check_consecutive(doctor, d, shift_type)
        # 周一有"恰好"约束，优先填满。
        # X/CT：bonus=5 强力把人吸到周一恰好。
        # MR：每天都"恰好6"，但周一已被陈绪珠占 1，初始 gap 比其他天少 1，
        # 完全不给 bonus 会让贪心偏离周一（周一最终凑不到 6）；给 +2 把 score 拉
        # 到与其他天接近，作为 tiebreaker，过强会把 MR 都堆到周一导致后续天不足 6。
        monday_bonus = 0
        if d == 0 and gap > 0:
            monday_bonus = 5 if shift_type != 'MR' else 2
        # 均衡软约束: 超过8人时递增惩罚(不硬性阻止,只降低优先级)
        balance_penalty = 0
        if d != 0 and counts[d][shift_type] >= 8:
            balance_penalty = (counts[d][shift_type] - 7) * 3
        # 被辅助人排MR时，优先有盯机覆盖的天
        dinji_bonus = 0
        if doctor and doctor.name in SENIORS and shift_type == 'MR':
            for aname in ASSISTANTS:
                if aname not in _doctors_ref:
                    continue
                adoc = _doctors_ref[aname]
                orig = adoc.original[d]
                if orig and re.match(r'MR\d+盯机', str(orig).strip()):
                    if can_pair_senior(aname, doctor.name):
                        dinji_bonus = 15
                        break
        candidates.append((gap - penalty + monday_bonus + dinji_bonus - balance_penalty, gap, d))

    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][2]


def has_mr_assist_available(senior_name, day_idx):
    """检查某天是否有MR辅助候选人可用(含MR盯机配对)"""
    # 路径1: 当天有纯辅助人MR盯机且能配对给该senior → 直接OK
    for name in ASSISTANTS:
        if name not in _doctors_ref:
            continue
        doc = _doctors_ref[name]
        orig = doc.original[day_idx]
        if orig and re.match(r'MR\d+盯机', str(orig).strip()):
            if can_pair_senior(name, senior_name):
                return True

    # 路径2: 从候选人列表找(考虑全周竞争)
    mr_candidates = get_candidates(senior_name, 'MR')
    usable_cands = []
    for cand in mr_candidates:
        if cand not in _doctors_ref:
            continue
        doc = _doctors_ref[cand]
        if cand in NO_MR_ASSIST:
            continue
        if doc.max_mr <= 0:
            continue
        usable_cands.append(cand)

    # 计算共享候选人被其他senior消耗的量
    shared_consumption = 0
    for sn in SENIORS:
        if sn == senior_name or sn not in _doctors_ref:
            continue
        sn_doc = _doctors_ref[sn]
        sn_cands = get_candidates(sn, 'MR')
        mr_days = 0
        for d in range(5):
            shift = sn_doc.schedule[d]
            if shift is None:
                orig = sn_doc.original[d]
                if orig and str(orig).strip() == 'MR':
                    shift = 'MR'
            if shift == 'MR':
                mr_days += 1
        shared = set(usable_cands) & set(sn_cands)
        if shared:
            shared_consumption += min(mr_days, len(shared))

    return len(usable_cands) - shared_consumption > 0


# 全局引用(在main中设置)
_doctors_ref = {}

def assign_quota(doctor, quota_dict, counts, blocked_days=None):
    if blocked_days is None:
        blocked_days = set()

    available = [d for d in range(5) if d not in blocked_days and doctor.original[d] is None]

    needed = []
    for st, cnt in quota_dict.items():
        if st == 'total':
            continue
        needed.extend([st] * cnt)

    if len(needed) > len(available):
        # 配额数 > 可用天数：必须裁剪。
        # MR 是 §二 的"恰好 6"硬约束，X/CT 只是"≥目标"软约束（不足可由其他人补）。
        # 个人 MR 配额一周只有 1 次，被裁掉就再也补不回来——天数受限的人尤其如此。
        # 因此 MR 在裁剪时**优先保留**：先按 (个人 MR 配额, 个人剩余 MR 容量,
        # 在可用且未被 mr_blocked_days 阻塞的天里 MR 还能塞进去的总 slot) 三者最小值
        # 决定保留多少个 MR；剩余位置按 X/CT 的当前缺口排序填充。
        n_avail = len(available)
        mr_blocked = doctor.constraints.get('mr_blocked_days', ())
        mr_in_quota = sum(1 for st in needed if st == 'MR')
        mr_capacity = max(0, doctor.max_mr - doctor.mr_count_this_week)
        mr_avail_slots = sum(
            max(0, get_target(d, 'MR') - counts[d]['MR'])
            for d in available if d not in mr_blocked
        )
        mr_to_keep = min(mr_in_quota, mr_capacity, mr_avail_slots, n_avail)

        non_mr = [st for st in needed if st != 'MR']
        scored = [
            (sum(max(0, get_target(d, st) - counts[d][st]) for d in available), st)
            for st in non_mr
        ]
        scored.sort(reverse=True)

        remaining = n_avail - mr_to_keep
        needed = ['MR'] * mr_to_keep + [s[1] for s in scored[:remaining]]

    used = set()
    # MR first
    for st in sorted(needed, key=lambda s: (0 if s == 'MR' else 1)):
        if st == 'MR' and doctor.mr_count_this_week >= doctor.max_mr:
            continue
        # 被辅助人排MR时,确保当天有辅助候选人可用
        extra_filter = None
        if doctor.name in SENIORS and st == 'MR':
            extra_filter = lambda d: has_mr_assist_available(doctor.name, d)
        day = find_best_day(st, available, used, counts, doctor, extra_filter)
        if day is not None:
            doctor.schedule[day] = st
            counts[day][st] += 1
            used.add(day)
            if st == 'MR':
                doctor.mr_count_this_week += 1


def assign_flexible(doctor, total, counts, blocked_days=None):
    if blocked_days is None:
        blocked_days = set()

    available = [d for d in range(5) if d not in blocked_days and doctor.original[d] is None]
    slots = min(total, len(available))
    used = set()

    is_hjb = doctor.constraints.get('hui_ji_bai', False)
    x_assigned = False

    for _ in range(slots):
        best_day, best_shift, best_score = None, None, -999
        for d in available:
            if d in used:
                continue
            # 回急白: 优先X(1个), 其次CT, 不排MR
            if is_hjb:
                shift_options = ['X', 'CT'] if not x_assigned else ['CT']
            else:
                shift_options = ['MR', 'CT', 'X']

            for st in shift_options:
                if counts[d][st] >= get_upper_limit(d, st):
                    continue
                if st == 'MR' and doctor.mr_count_this_week >= doctor.max_mr:
                    continue
                gap = get_target(d, st) - counts[d][st]
                penalty = check_consecutive(doctor, d, st)
                bonus = 5 if (is_hjb and st == 'X' and not x_assigned) else 0
                # 均衡软约束
                balance_penalty = 0
                if d != 0 and counts[d][st] >= 8:
                    balance_penalty = (counts[d][st] - 7) * 3
                score = gap - penalty + bonus - balance_penalty
                if score > best_score:
                    best_score = score
                    best_shift = st
                    best_day = d
        if best_day is not None:
            doctor.schedule[best_day] = best_shift
            counts[best_day][best_shift] += 1
            used.add(best_day)
            if best_shift == 'MR':
                doctor.mr_count_this_week += 1
            if best_shift == 'X' and is_hjb:
                x_assigned = True


# ============================================================
# Phases
# ============================================================
def phase1(doctors, counts):
    d = doctors['闫东']
    d.schedule[0] = '主班'; d.schedule[1] = '主班'
    d.schedule[2] = 'X'; counts[2]['X'] += 1
    d.schedule[3] = 'X'; counts[3]['X'] += 1
    d.schedule[4] = 'X'; counts[4]['X'] += 1
    d.needs_scheduling = False

    d = doctors['陈绪珠']
    d.schedule[0] = 'MR'; counts[0]['MR'] += 1
    d.needs_scheduling = False; d.mr_count_this_week = 1

    d = doctors['程晓光']
    d.schedule[0] = 'X'; counts[0]['X'] += 1
    d.schedule[1] = 'X'; counts[1]['X'] += 1
    d.needs_scheduling = False

    # 尤玉华: §九 排班固定不动 (休/休/休/X/CT)。col2='星期四X，星期五CT' 不被通用
    # 正则解析；为保证她周四 X、周五 CT 真正落到 schedule 上(而非依赖 original 已填),
    # 在 phase1 直接写。本周 MR=0。
    d = doctors['尤玉华']
    d.schedule[3] = 'X'; counts[3]['X'] += 1
    d.schedule[4] = 'CT'; counts[4]['CT'] += 1
    d.needs_scheduling = False
    d.mr_count_this_week = 0

    # 被辅助人中有MR配额的，按MR辅助候选人受限程度排序（最受限的先排）
    seniors_with_mr = []
    for name in SENIORS:
        if name not in doctors or not doctors[name].needs_scheduling:
            continue
        d = doctors[name]
        if d.quota and 'MR' in d.quota and d.quota['MR'] > 0:
            # 计算实际可用MR辅助候选人数
            cands = get_candidates(name, 'MR')
            usable = sum(1 for c in cands if c in doctors and c not in NO_MR_ASSIST
                        and doctors[c].max_mr > 0)
            seniors_with_mr.append((usable, name))
    seniors_with_mr.sort()  # 可用候选越少越优先

    for _, name in seniors_with_mr:
        d = doctors[name]
        blocked = {i for i in range(5) if d.original[i] is not None}
        assign_quota(d, {k: v for k, v in d.quota.items()}, counts, blocked)
        d.needs_scheduling = False

    # 翁磊: 如果还没被seniors_with_mr处理(没有MR配额的情况), 这里处理
    d = doctors['翁磊']
    if d.needs_scheduling:
        # 没有 §九 hardcode 后, 配额完全由 col2 解析；这里只在 phase1 没拿到 MR 配额
        # (即 col2 不含 MR) 时兜底，按 col2 当前 quota 走 assign_quota.
        assign_quota(d, {k: v for k, v in d.quota.items()}, counts)
        d.needs_scheduling = False

    d = doctors['赵涛']
    # §九 赵涛"每周恰好2X，时间随意"。本周周五 5 名全效辅助人参会，senior 辅助资源
    # 耗尽——让赵涛主动避开周五，减少 1 个 senior 周五 X 班，使周五 senior 辅助缺口
    # 从 3 降到 2。这是算法选项, 不违 §九 "时间随意"。
    assign_quota(d, {'X': 2}, counts, blocked_days={4})
    d.needs_scheduling = False

    d = doctors['白荣杰']
    # §九 白荣杰 没规定具体日期；同上原因 (周五辅助资源耗尽)，让他避开周五。
    # col2='3X' 共 3 班，在周一-四 4 天里选 3 天，并不违反任何约束。
    # 收益：白荣杰移出周五后，杨柳的周五 X 辅助 (pin) 可以让给闫东，使周五 senior
    # 辅助缺口从 2 降到 1 (尤玉华 CT 仍无辅助)。
    assign_quota(d, {k: v for k, v in d.quota.items()}, counts, blocked_days={4})
    d.needs_scheduling = False

    d = doctors['王玲']
    mr_gap_thu = 6 - counts[3]['MR']
    mr_gap_fri = 6 - counts[4]['MR']
    if mr_gap_thu >= mr_gap_fri:
        d.schedule[3] = 'MR'; counts[3]['MR'] += 1
        d.schedule[4] = 'X'; counts[4]['X'] += 1
    else:
        d.schedule[3] = 'X'; counts[3]['X'] += 1
        d.schedule[4] = 'MR'; counts[4]['MR'] += 1
    d.needs_scheduling = False; d.mr_count_this_week = 1


def phase2(doctors, counts):
    """追加排班 - 处理 pin_day 约束 (§四 列3 写'在周X/在星期X'的人)。
    - 普通医生：把 1 班放到 pin_day。
        cell 是 appendable（胃肠/开药+/输卵管）或纯介入 → 设 append_day/append_shift，
            phase7 渲染为 '<cell>+<shift>'。
        cell 空 → 直接写 schedule[pin_day]=shift。
        cell 是其他固定班/复合介入 → 跳过 pin（无处可放）并打印警告。
        班种：回急白者优先 X（§5.4），否则按当天 X/CT 缺口选（不能 MR，§5.1）。
    - 辅助人：设 append_day/is_pinned_assist，phase5/phase6 自然把 pin_day 优先安排
        （依赖 is_asst_available 中已有的 pinned 检查）。
    """
    # 普通医生 pin_day
    for name, d in doctors.items():
        if d.category == 'assistant':
            continue
        pin_day = d.constraints.get('pin_day')
        if pin_day is None or not d.needs_scheduling:
            continue

        if d.constraints.get('hui_ji_bai'):
            if counts[pin_day]['X'] < get_upper_limit(pin_day, 'X'):
                shift = 'X'
            else:
                shift = 'CT'
        else:
            shift = pick_best_shift_no_mr(pin_day, counts)

        orig = d.original[pin_day]
        if is_empty(orig):
            d.schedule[pin_day] = shift
            counts[pin_day][shift] += 1
        elif is_appendable(orig) or is_pure_jieru(orig):
            d.constraints['append_day'] = pin_day
            d.constraints['append_shift'] = shift
            counts[pin_day][shift] += 1
        else:
            print(f"  ⚠️ {name} pin_day={DAYS[pin_day]} 格子 {orig!r} 不可追加, 跳过 pin")
            continue

        if 'total' in d.quota:
            d.quota['total'] = max(0, d.quota['total'] - 1)
            if d.quota['total'] <= 0:
                d.needs_scheduling = False

    # 辅助人 pin_day → 让 phase5/phase6 把这一天的辅助优先留给本人
    for name, d in doctors.items():
        if d.category != 'assistant':
            continue
        pin_day = d.constraints.get('pin_day')
        if pin_day is None:
            continue
        d.constraints['append_day'] = pin_day
        d.constraints['is_pinned_assist'] = True


def pick_best_shift_no_mr(day_idx, counts):
    """选X或CT(不能MR)"""
    gaps = {}
    for st in ['X', 'CT']:
        if counts[day_idx][st] >= get_upper_limit(day_idx, st):
            continue
        gaps[st] = get_target(day_idx, st) - counts[day_idx][st]
    if not gaps:
        return 'X'
    return max(gaps, key=gaps.get)


def phase3(doctors, counts):
    to_assign = []
    for name, d in doctors.items():
        if not d.needs_scheduling or d.category == 'assistant':
            continue
        if 'total' not in d.quota and d.quota:
            to_assign.append(d)
    to_assign.sort(key=lambda d: sum(d.quota.values()), reverse=True)
    for d in to_assign:
        assign_quota(d, {k: v for k, v in d.quota.items()}, counts)
        d.needs_scheduling = False


def phase4(doctors, counts):
    to_assign = []
    for name, d in doctors.items():
        if not d.needs_scheduling or d.category == 'assistant':
            continue
        if 'total' in d.quota and d.quota['total'] > 0:
            to_assign.append(d)
    to_assign.sort(key=lambda d: d.quota.get('total', 0), reverse=True)
    for d in to_assign:
        assign_flexible(d, d.quota['total'], counts)
        d.needs_scheduling = False


# ============================================================
# Phase 5 & 6: 辅助(含半效规则)
# ============================================================
def can_pair_senior(asst_name, senior_name):
    """检查盯机辅助人能否覆盖某senior"""
    if senior_name == '陈绪珠':
        return asst_name in ['季爱华', '赵海清', '林晨', '王业学']
    if senior_name == '尤玉华' and asst_name == '师欣瑶':
        return False
    if senior_name in ('程晓光', '翁磊') and asst_name == '杨柳':
        return False
    return True


def phase5(doctors, counts):
    """第一阶段辅助"""
    assignments = []  # [(day, asst_name, senior_name, shift_type)]

    # 预计算: 纯辅助人MR盯机覆盖了哪些(day, senior)
    mr_dinji_coverage = set()
    for name in ASSISTANTS:
        if name not in doctors:
            continue
        doc = doctors[name]
        for day_idx in range(5):
            orig = doc.original[day_idx]
            if orig and re.match(r'MR\d+盯机', str(orig).strip()):
                for sn in SENIORS:
                    if sn not in doctors:
                        continue
                    sn_doc = doctors[sn]
                    sn_shift = sn_doc.schedule[day_idx]
                    if sn_shift is None:
                        sn_orig = sn_doc.original[day_idx]
                        if sn_orig and str(sn_orig).strip() == 'MR':
                            sn_shift = 'MR'
                    if sn_shift == 'MR':
                        if can_pair_senior(name, sn):
                            mr_dinji_coverage.add((day_idx, sn))
                            break

    for day_idx in range(5):
        needs = []
        for sn in SENIORS:
            if sn not in doctors: continue
            d = doctors[sn]
            shift = d.schedule[day_idx]
            if shift is None:
                orig = d.original[day_idx]
                if orig and str(orig).strip() in ('X', 'CT', 'MR'):
                    shift = str(orig).strip()
            if shift in ('X', 'CT', 'MR'):
                # MR且被盯机覆盖 → 跳过
                if shift == 'MR' and (day_idx, sn) in mr_dinji_coverage:
                    continue
                needs.append((sn, shift))

        # 排序needs: 实际可用辅助候选人越少的越优先处理
        def restriction_score(item):
            sn, st = item
            cands = get_candidates(sn, st)
            # 计算实际可用候选人数
            available_count = 0
            for c in cands:
                if c not in doctors: continue
                cd = doctors[c]
                if st == 'MR' and c in NO_MR_ASSIST: continue
                if st == 'MR' and cd.mr_count_this_week >= cd.max_mr: continue
                orig = cd.original[day_idx]
                if orig is None or is_pure_jieru(orig):
                    if not any(d == day_idx and asst == c for (d, asst, _, _) in assignments):
                        available_count += 1
            return available_count
        # 排序: 主键 = MAY_LACK_ASSIST 中的 (senior, day) 排在最后 (允许无辅助);
        #       次键 = restriction_score (候选少的优先).
        # MAY_LACK_ASSIST: 用户明确允许在辅助资源耗尽时跳过的 senior 班次.
        # 当前仅闫东周五——周五 5 个全效辅助参会, 闫东 §九 周三四五X 固定无法挪开,
        # 而尤玉华 §九 周五CT 同样固定但用户要求她必须有辅助; 给闫东周五打低优先级,
        # 让 杨柳 (周五唯一可调辅助) 先满足尤玉华.
        MAY_LACK_ASSIST = {('闫东', 4)}
        needs.sort(key=lambda item: (
            1 if (item[0], day_idx) in MAY_LACK_ASSIST else 0,
            restriction_score(item)
        ))

        for senior_name, shift_type in needs:
            success = assign_one_assist(
                senior_name, shift_type, day_idx, doctors, assignments
            )
            if not success:
                print(f"  ⚠️ {DAYS[day_idx]} {senior_name}({shift_type}) 无法分配辅助")

    return assignments


def assign_one_assist(senior_name, shift_type, day_idx, doctors, assignments):
    """为一个班次分配辅助(全效1人或半效2人, 兜底半效1人)"""

    def score_candidate(cand_name):
        """评估辅助候选人: 考虑多样性和回急白优先X"""
        cand = doctors[cand_name]
        existing_types = []
        for (d, asst, _, st) in assignments:
            if asst == cand_name:
                existing_types.append(st)
        same_count = existing_types.count(shift_type)
        # 强化多样性: 已做过相同类型越多惩罚越重
        diversity_penalty = same_count * 10
        # 连续天检查
        consec_penalty = 0
        if day_idx > 0:
            for (d, asst, _, st) in assignments:
                if d == day_idx - 1 and asst == cand_name and st == shift_type:
                    consec_penalty = 5
        if day_idx < 4:
            for (d, asst, _, st) in assignments:
                if d == day_idx + 1 and asst == cand_name and st == shift_type:
                    consec_penalty = 5
        # 回急白的辅助人: 优先分配给X, 非X给大惩罚
        hjb_penalty = 0
        if cand.constraints.get('hui_ji_bai') and shift_type != 'X':
            hjb_penalty = 20
        return -(diversity_penalty + consec_penalty + hjb_penalty)

    # 先尝试全效(按多样性排序)
    full_cand = get_candidates_by_type(senior_name, shift_type, day_idx, doctors, assignments, 'full')
    valid_full = [(c, score_candidate(c)) for c in full_cand
                  if can_assign_assist(c, shift_type, day_idx, doctors, assignments, senior_name)]
    valid_full.sort(key=lambda x: x[1], reverse=True)

    for cand_name, _ in valid_full:
        assignments.append((day_idx, cand_name, senior_name, shift_type))
        consume_asst_quota(doctors[cand_name], shift_type)
        return True

    # 再尝试半效(需要2人)
    half_cand = get_candidates_by_type(senior_name, shift_type, day_idx, doctors, assignments, 'half')
    available_halfs = [(c, score_candidate(c)) for c in half_cand
                       if can_assign_assist(c, shift_type, day_idx, doctors, assignments, senior_name)]
    available_halfs.sort(key=lambda x: x[1], reverse=True)

    if len(available_halfs) >= 2:
        h1, h2 = available_halfs[0][0], available_halfs[1][0]
        assignments.append((day_idx, h1, senior_name, shift_type))
        assignments.append((day_idx, h2, senior_name, shift_type))
        consume_asst_quota(doctors[h1], shift_type)
        consume_asst_quota(doctors[h2], shift_type)
        return True

    # 兜底: 1个半效人独立辅助(一周最多1次)
    if len(available_halfs) >= 1:
        h1 = available_halfs[0][0]
        # 检查该半效人本周是否已独立辅助过
        solo_count = sum(1 for (d, asst, sn, st) in assignments
                        if asst == h1 and asst in HALF_EFFICIENCY
                        # 检查是否是独立的(同天同senior没有另一个半效partner)
                        and not any(d2 == d and asst2 != asst and asst2 in HALF_EFFICIENCY and sn2 == sn
                                   for (d2, asst2, sn2, st2) in assignments))
        if solo_count < 1:
            assignments.append((day_idx, h1, senior_name, shift_type))
            consume_asst_quota(doctors[h1], shift_type)
            return True

    return False


def get_candidates_by_type(senior_name, shift_type, day_idx, doctors, assignments, eff_type):
    """获取候选辅助人(按效率类型和优先级)"""
    base = get_candidates(senior_name, shift_type)
    if eff_type == 'full':
        return [c for c in base if c in FULL_EFFICIENCY]
    else:
        return [c for c in base if c in HALF_EFFICIENCY]


def get_candidates(senior_name, shift_type):
    if senior_name == '陈绪珠':
        return ['季爱华', '赵海清', '林晨', '王业学']
    if senior_name == '白荣杰':
        if shift_type == 'MR':
            # 只从这4人中选，不考虑其他人
            return ['陈垒', '季爱华', '陈建安', '赵海清']
        else:
            return ['王苹'] + [a for a in ASSISTANTS if a != '王苹']
    return list(ASSISTANTS)


def can_assign_assist(cand_name, shift_type, day_idx, doctors, assignments, senior_name=''):
    """检查辅助人是否可被分配"""
    cand = doctors[cand_name]

    # MR限制
    if shift_type == 'MR' and cand_name in NO_MR_ASSIST:
        return False
    # 尤玉华不用师欣瑶
    if senior_name == '尤玉华' and cand_name == '师欣瑶':
        return False
    # 杨柳不给程晓光和翁磊当辅助
    if cand_name == '杨柳' and senior_name in ('程晓光', '翁磊'):
        return False
    # MR周限制
    if shift_type == 'MR' and cand.mr_count_this_week >= cand.max_mr:
        return False
    # 当天已被分配
    if any(d == day_idx and asst == cand_name for (d, asst, _, _) in assignments):
        return False
    # 当天可用
    if not is_asst_available(cand, day_idx, assignments):
        return False
    # 有配额
    if not has_quota(cand, shift_type):
        return False
    return True


def is_asst_available(doctor, day_idx, existing):
    for (d, asst, _, _) in existing:
        if d == day_idx and asst == doctor.name:
            return False

    # pinned assist检查
    if doctor.constraints.get('is_pinned_assist'):
        pin = doctor.constraints.get('append_day')
        if pin is not None and day_idx != pin:
            remaining = get_remaining(doctor)
            pinned_consumed = any(d == pin and asst == doctor.name for (d, asst, _, _) in existing)
            if not pinned_consumed:
                if remaining - 1 <= 0:
                    return False

    orig = doctor.original[day_idx]
    sched = doctor.schedule[day_idx]

    if doctor.constraints.get('append_day') == day_idx:
        return True
    if is_empty(orig) and sched is None:
        return True
    if is_pure_jieru(orig) and sched is None:
        return True
    return False


def has_quota(doctor, shift_type):
    q = doctor.quota
    if not q: return False
    if 'total' in q: return q['total'] > 0
    if shift_type in q: return q[shift_type] > 0
    return False


def consume_asst_quota(doctor, shift_type):
    q = doctor.quota
    if 'total' in q:
        q['total'] = max(0, q['total'] - 1)
    elif shift_type in q:
        q[shift_type] = max(0, q[shift_type] - 1)
    if shift_type == 'MR':
        doctor.mr_count_this_week += 1


def get_remaining(doctor):
    q = doctor.quota
    if not q: return 0
    if 'total' in q: return q['total']
    return sum(v for v in q.values() if isinstance(v, int) and v > 0)


def phase6(doctors, stage1):
    """第二阶段辅助"""
    stage2 = []
    all_assignments = list(stage1)

    # 记录已有辅助覆盖的(day, receiver)组合
    covered = set()
    for (day_idx, _, senior_name, _) in stage1:
        covered.add((day_idx, senior_name))

    # MR盯机覆盖: 盯机本身就是辅助,不需要额外辅助
    for name, doc in doctors.items():
        for day_idx in range(5):
            orig = doc.original[day_idx]
            if orig and re.match(r'MR\d+盯机', str(orig).strip()):
                if name not in ASSISTANTS:
                    # 非辅助人盯机: 自己不需要辅助
                    covered.add((day_idx, name))
                else:
                    # 纯辅助人盯机: 找能配对的医生(用can_pair_senior规则)
                    for doc_name, doc2 in doctors.items():
                        if doc2.schedule[day_idx] == 'MR' and can_pair_senior(name, doc_name):
                            covered.add((day_idx, doc_name))
                            break
                        # 也检查original中的MR
                        if doc2.original[day_idx] and str(doc2.original[day_idx]).strip() == 'MR':
                            if can_pair_senior(name, doc_name):
                                covered.add((day_idx, doc_name))
                                break

    receivers = []
    for name in ASSIST_PRIORITY_HIGHEST:
        if name in doctors: receivers.append(name)
    for name in ASSIST_PRIORITY_SECOND:
        if name in doctors: receivers.append(name)
    for name, d in doctors.items():
        if d.category == 'normal' and name not in receivers and name not in NEVER_RECEIVE_ASSIST:
            receivers.append(name)

    # 均衡计数器(用于ASSIST_PRIORITY_SECOND的均衡)
    recv_assist_count = {n: 0 for n in receivers}

    def get_sorted_receivers():
        """动态排序: 裴京哲第一, 然后5人按辅助次数升序, 最后其余"""
        highest = [n for n in ASSIST_PRIORITY_HIGHEST if n in doctors]
        second = sorted([n for n in ASSIST_PRIORITY_SECOND if n in doctors],
                       key=lambda n: recv_assist_count.get(n, 0))
        others = [n for n in receivers if n not in highest and n not in second]
        return highest + second + others

    # 先处理全效辅助人
    for asst_doc in [doctors[n] for n in FULL_EFFICIENCY if n in doctors and get_remaining(doctors[n]) > 0]:
        while get_remaining(asst_doc) > 0:
            matched = False
            needed_types = get_needed_types(asst_doc)
            for recv_name in get_sorted_receivers():
                if matched: break
                recv = doctors[recv_name]
                for day_idx in range(5):
                    if matched: break
                    if (day_idx, recv_name) in covered:
                        continue
                    shift = recv.schedule[day_idx]
                    if shift is None:
                        orig = recv.original[day_idx]
                        if orig and str(orig).strip() in ('X', 'CT', 'MR'):
                            shift = str(orig).strip()
                    if shift not in ('X', 'CT', 'MR'):
                        continue
                    if needed_types and shift not in needed_types:
                        continue
                    if shift == 'MR' and asst_doc.name in NO_MR_ASSIST:
                        continue
                    if shift == 'MR' and asst_doc.mr_count_this_week >= asst_doc.max_mr:
                        continue
                    if not is_asst_available(asst_doc, day_idx, all_assignments):
                        continue

                    stage2.append((day_idx, asst_doc.name, recv_name, shift))
                    all_assignments.append((day_idx, asst_doc.name, recv_name, shift))
                    consume_asst_quota(asst_doc, shift)
                    covered.add((day_idx, recv_name))
                    recv_assist_count[recv_name] = recv_assist_count.get(recv_name, 0) + 1
                    matched = True
            if not matched:
                break

    # 再处理半效辅助人(需要成对, 找不到partner则兜底单独辅助)
    half_with_quota = [doctors[n] for n in HALF_EFFICIENCY if n in doctors and get_remaining(doctors[n]) > 0]
    # 按剩余配额降序(配额多的优先处理, 避免被饿死)
    half_with_quota.sort(key=lambda d: get_remaining(d), reverse=True)

    # 用索引遍历，避免重复配对
    i = 0
    while i < len(half_with_quota):
        asst_doc = half_with_quota[i]
        if get_remaining(asst_doc) <= 0:
            i += 1
            continue

        matched = False
        needed_types = get_needed_types(asst_doc)

        for recv_name in get_sorted_receivers():
            if matched: break
            recv = doctors[recv_name]
            for day_idx in range(5):
                if matched: break
                if (day_idx, recv_name) in covered:
                    continue
                shift = recv.schedule[day_idx]
                if shift is None:
                    orig = recv.original[day_idx]
                    if orig and str(orig).strip() in ('X', 'CT', 'MR'):
                        shift = str(orig).strip()
                if shift not in ('X', 'CT', 'MR'):
                    continue
                if needed_types and shift not in needed_types:
                    continue
                if shift == 'MR' and asst_doc.name in NO_MR_ASSIST:
                    continue
                if shift == 'MR' and asst_doc.mr_count_this_week >= asst_doc.max_mr:
                    continue
                if not is_asst_available(asst_doc, day_idx, all_assignments):
                    continue

                # 找半效partner
                partner = None
                for other in half_with_quota:
                    if other.name == asst_doc.name: continue
                    if get_remaining(other) <= 0: continue
                    if shift == 'MR' and other.name in NO_MR_ASSIST: continue
                    if shift == 'MR' and other.mr_count_this_week >= other.max_mr: continue
                    nt = get_needed_types(other)
                    if nt and shift not in nt: continue
                    if not is_asst_available(other, day_idx, all_assignments): continue
                    partner = other
                    break

                if partner:
                    stage2.append((day_idx, asst_doc.name, recv_name, shift))
                    stage2.append((day_idx, partner.name, recv_name, shift))
                    all_assignments.append((day_idx, asst_doc.name, recv_name, shift))
                    all_assignments.append((day_idx, partner.name, recv_name, shift))
                    consume_asst_quota(asst_doc, shift)
                    consume_asst_quota(partner, shift)
                    covered.add((day_idx, recv_name))
                    recv_assist_count[recv_name] = recv_assist_count.get(recv_name, 0) + 1
                    matched = True
                else:
                    # 兜底: 半效人单独辅助(一周最多1次)
                    solo_count = sum(1 for (d2, asst2, sn2, st2) in all_assignments
                                    if asst2 == asst_doc.name and asst2 in HALF_EFFICIENCY
                                    and not any(d3 == d2 and asst3 != asst2 and asst3 in HALF_EFFICIENCY and sn3 == sn2
                                               for (d3, asst3, sn3, st3) in all_assignments))
                    if solo_count < 1:
                        stage2.append((day_idx, asst_doc.name, recv_name, shift))
                        all_assignments.append((day_idx, asst_doc.name, recv_name, shift))
                        consume_asst_quota(asst_doc, shift)
                        covered.add((day_idx, recv_name))
                        recv_assist_count[recv_name] = recv_assist_count.get(recv_name, 0) + 1
                        matched = True

        if not matched:
            i += 1
        # Don't increment i if matched - check if still has quota

    return stage2


def get_needed_types(doctor):
    q = doctor.quota
    if 'total' in q: return []
    return [st for st, cnt in q.items() if isinstance(cnt, int) and cnt > 0]


# ============================================================
# Phase 6.5: 后处理优化 - 消除连续同班种
# ============================================================
def phase6_5_optimize(doctors, counts):
    """通过交换消除连续同班种"""
    improved = True
    iterations = 0
    while improved and iterations < 20:
        improved = False
        iterations += 1
        for i_name, doc_i in doctors.items():
            if doc_i.category == 'assistant':
                continue
            for day_idx in range(1, 5):
                shift_i = doc_i.schedule[day_idx]
                if shift_i not in ('X', 'CT', 'MR'):
                    continue
                prev_i = doc_i.schedule[day_idx - 1] if day_idx > 0 else None
                if prev_i != shift_i:
                    continue
                # doc_i 在 day_idx 和 day_idx-1 有连续同班种
                # 找另一个人在 day_idx 有不同班种, 互换
                for j_name, doc_j in doctors.items():
                    if j_name == i_name or doc_j.category == 'assistant':
                        continue
                    shift_j = doc_j.schedule[day_idx]
                    if shift_j not in ('X', 'CT', 'MR'):
                        continue
                    if shift_j == shift_i:
                        continue
                    # 检查互换是否合法
                    if not can_swap(doc_i, doc_j, day_idx, shift_i, shift_j, counts):
                        continue
                    # 检查互换后doc_j不会产生新的连续
                    prev_j = doc_j.schedule[day_idx - 1] if day_idx > 0 else None
                    next_j = doc_j.schedule[day_idx + 1] if day_idx < 4 else None
                    if prev_j == shift_i or next_j == shift_i:
                        continue  # 会给j造成连续
                    # 执行互换
                    doc_i.schedule[day_idx] = shift_j
                    doc_j.schedule[day_idx] = shift_i
                    # headcount不变(同天互换)
                    improved = True
                    break
    return iterations


def fix_mr_shortfall(doctors, counts):
    """补 MR < 6 的天。
    思路：贪心+裁剪让某些天的 MR 凑不齐 6（§二 硬约束）。这里做"两人三班"修复——
      P：当天(short_d) 排了 X/CT 且仍有 MR 容量 → 换成 MR
      Q：另一天(n)排了同班种 X/CT、那天该班种数 > 目标、且 Q 在 short_d 没班 → 把 Q
         的 X/CT 从 n 挪到 short_d，顶上 P 让出的格子
    净效果：short_d 该 X/CT 数不变、short_d MR +1、n 该 X/CT 数 -1（仍 ≥ target）。
    每次循环只补 1 个 short slot，最多 20 次防止无限循环。

    锁住的人不参与（rules.md §九 中"排班固定不动"或"周X 固定"的人）：
      程晓光 周一二 X、闫东 周一二主班+周三四五X、陈绪珠 周一 MR+其余主班、
      尤玉华 全周固定、王玲 周四五 1MR+1X。
    """
    LOCKED = {'程晓光', '闫东', '陈绪珠', '尤玉华', '王玲'}
    iterations = 0
    while iterations < 20:
        iterations += 1
        short_days = [d for d in range(5) if counts[d]['MR'] < get_target(d, 'MR')]
        if not short_days:
            break
        moved = False
        for short_d in short_days:
            for p_name, p in doctors.items():
                if p_name in LOCKED:
                    continue
                if p.category == 'assistant':
                    continue
                if p.constraints.get('hui_ji_bai'):
                    continue
                if short_d in p.constraints.get('mr_blocked_days', ()):
                    continue
                if p.mr_count_this_week >= p.max_mr:
                    continue
                p_old = p.schedule[short_d]
                if p_old not in ('X', 'CT'):
                    continue
                # 寻找 Q 顶替 P 在 short_d 的 X/CT
                for q_name, q in doctors.items():
                    if q is p or q_name in LOCKED or q.category == 'assistant':
                        continue
                    if q.original[short_d] is not None:
                        continue
                    if q.schedule[short_d] is not None:
                        continue
                    # Q 必须在某 n!=short_d 排了 p_old，且那天减 1 仍 ≥ target，且 n!=0（周一恰好）
                    n_pick = None
                    for n in range(5):
                        if n == short_d or n == 0:
                            continue
                        if q.schedule[n] != p_old:
                            continue
                        if counts[n][p_old] - 1 < get_target(n, p_old):
                            continue
                        n_pick = n
                        break
                    if n_pick is None:
                        continue
                    # 应用 swap
                    p.schedule[short_d] = 'MR'
                    counts[short_d][p_old] -= 1
                    counts[short_d]['MR'] += 1
                    p.mr_count_this_week += 1
                    q.schedule[n_pick] = None
                    counts[n_pick][p_old] -= 1
                    q.schedule[short_d] = p_old
                    counts[short_d][p_old] += 1
                    moved = True
                    break
                if moved:
                    break
            if moved:
                break
        if not moved:
            break
    return iterations


def can_swap(doc_i, doc_j, day_idx, shift_i, shift_j, counts):
    """检查两人在同一天互换班种是否合法"""
    # 有明确配额的人不能互换(会破坏配额约束)
    if doc_i.quota and 'total' not in doc_i.quota:
        return False
    if doc_j.quota and 'total' not in doc_j.quota:
        return False

    # MR周限制
    if shift_j == 'MR':
        if doc_i.mr_count_this_week >= doc_i.max_mr and shift_i != 'MR':
            return False
    if shift_i == 'MR':
        if doc_j.mr_count_this_week >= doc_j.max_mr and shift_j != 'MR':
            return False

    # 回急白的人不能接收MR
    if shift_j == 'MR' and doc_i.constraints.get('hui_ji_bai'):
        return False
    if shift_i == 'MR' and doc_j.constraints.get('hui_ji_bai'):
        return False

    # 个人 MR 禁排日（rules.md §九，例如李新民周一不排MR）
    if shift_j == 'MR' and day_idx in doc_i.constraints.get('mr_blocked_days', ()):
        return False
    if shift_i == 'MR' and day_idx in doc_j.constraints.get('mr_blocked_days', ()):
        return False

    # 被辅助人不互换(辅助配对会失效)
    if doc_i.name in SENIORS or doc_j.name in SENIORS:
        return False

    return True


# ============================================================
# Phase 7: 编号与输出
# ============================================================
def phase7(doctors, df, stage1, stage2, filepath):
    all_assists = stage1 + stage2

    # Step 1: MR编号从4开始(1-3是盯机), X/CT从1开始
    daily_people = {d: {'X': [], 'CT': [], 'MR': []} for d in range(5)}

    for i in range(2, len(df)):
        name = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ''
        if not name or name == 'nan' or name == '马毅民' or name not in doctors:
            continue
        d = doctors[name]
        for day_idx in range(5):
            shift = None
            if d.schedule[day_idx] in ('X', 'CT', 'MR'):
                shift = d.schedule[day_idx]
            elif d.original[day_idx] and str(d.original[day_idx]).strip() in ('X', 'CT', 'MR'):
                shift = str(d.original[day_idx]).strip()
            if d.constraints.get('append_day') == day_idx and d.category != 'assistant':
                shift = d.constraints.get('append_shift')
            if shift in ('X', 'CT', 'MR'):
                daily_people[day_idx][shift].append(name)

    # Step 2: 编号
    # 纯辅助人MR盯机需要配对: 找出哪些天有纯辅助人盯机, 分配对应编号
    # 配对需遵循辅助人规则(陈绪珠只接受特定辅助人等)
    asst_dinji = {}  # {(day_idx, mr_num): asst_name}
    for i in range(2, len(df)):
        name = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ''
        if name not in ASSISTANTS:
            continue
        for day_idx in range(5):
            val = str(df.iloc[i, DAY_COLS[day_idx]]).strip() if pd.notna(df.iloc[i, DAY_COLS[day_idx]]) else ''
            m = re.match(r'MR(\d+)盯机', val)
            if m:
                mr_num = int(m.group(1))
                asst_dinji[(day_idx, mr_num)] = name

    def can_pair_dinji(asst_name, doctor_name):
        """检查辅助人盯机能否配对给某医生(遵循辅助规则)"""
        # 陈绪珠只接受特定辅助人
        if doctor_name == '陈绪珠':
            allowed = ['季爱华', '赵海清', '林晨', '王业学']
            return asst_name in allowed
        # 尤玉华不接受师欣瑶
        if doctor_name == '尤玉华' and asst_name == '师欣瑶':
            return False
        # 白荣杰MR辅助优先特定人(但不限制)
        # 其他人无限制
        return True

    numbering = {}
    for day_idx in range(5):
        for st in ['X', 'CT']:
            for idx, name in enumerate(daily_people[day_idx][st], 1):
                numbering[(day_idx, name)] = f"{st}{idx}"

        # MR编号: 先分配需要配对的MR1/2/3给对应的医生
        mr_people = daily_people[day_idx]['MR']
        reserved_nums = {}  # {mr_num: person_name}

        for (d, mr_num), asst_name in asst_dinji.items():
            if d == day_idx:
                # 找一个符合辅助规则的配对医生
                for person in mr_people:
                    if person in reserved_nums.values():
                        continue
                    if can_pair_dinji(asst_name, person):
                        reserved_nums[mr_num] = person
                        break

        # 给reserved的人分配对应编号
        used_people = set()
        for mr_num, person in reserved_nums.items():
            numbering[(day_idx, person)] = f"MR{mr_num}"
            used_people.add(person)

        # 剩余MR人员从4开始编号
        next_num = 4
        for person in mr_people:
            if person in used_people:
                continue
            numbering[(day_idx, person)] = f"MR{next_num}"
            next_num += 1

    # Step 3: 辅助显示
    assist_display = {}
    for (day_idx, asst_name, senior_name, shift_type) in all_assists:
        senior_num = numbering.get((day_idx, senior_name))
        if senior_num:
            assist_display.setdefault((day_idx, asst_name), senior_num + '辅助')
        else:
            assist_display.setdefault((day_idx, asst_name), f"{shift_type}?辅助")

    # Step 4: 写入
    wb = load_workbook(filepath)
    ws = wb.active

    for name, d in doctors.items():
        row = d.row_idx + 1
        for day_idx in range(5):
            col = DAY_COLS[day_idx] + 1

            if (day_idx, name) in assist_display:
                display = assist_display[(day_idx, name)]
                orig = d.original[day_idx]
                if orig and is_pure_jieru(orig):
                    display = f"介入+{display}"
                elif orig and is_appendable(orig):
                    display = f"{str(orig).strip()}+{display}"
                elif orig and str(orig).strip().startswith('介入+'):
                    # 介入+追加班 复合 cell (例: 杨柳 周五='介入+开药+'): 仍视为可在末尾
                    # 追加辅助显示, 避免覆盖原 cell 内容。
                    display = f"{str(orig).strip()}+{display}"
                ws.cell(row=row, column=col, value=display)
                continue

            if d.constraints.get('append_day') == day_idx and d.category != 'assistant':
                orig = d.original[day_idx]
                num_str = numbering.get((day_idx, name), d.constraints.get('append_shift', ''))
                ws.cell(row=row, column=col, value=f"{str(orig).strip()}+{num_str}")
                continue

            shift = d.schedule[day_idx]
            if shift in ('X', 'CT', 'MR'):
                num_str = numbering.get((day_idx, name), shift)
                ws.cell(row=row, column=col, value=num_str)
            elif shift == '主班':
                ws.cell(row=row, column=col, value='主班')
            elif d.original[day_idx] and str(d.original[day_idx]).strip() in ('X', 'CT', 'MR'):
                num_str = numbering.get((day_idx, name), str(d.original[day_idx]).strip())
                ws.cell(row=row, column=col, value=num_str)

    out = filepath.replace('.xlsx', '_result.xlsx')
    wb.save(out)
    print(f"  ✅ 保存至: {out}")
    return out


# ============================================================
# 验证
# ============================================================
def validate(counts):
    print("\n=== 验证 ===")
    ok = True
    for day_idx in range(5):
        x_t, ct_t, mr_t = DAILY_TARGETS[DAYS[day_idx]]
        x, ct, mr = counts[day_idx]['X'], counts[day_idx]['CT'], counts[day_idx]['MR']
        issues = []
        if day_idx == 0:
            if x != x_t: issues.append(f"X={x}(需{x_t})")
            if ct != ct_t: issues.append(f"CT={ct}(需{ct_t})")
        else:
            if x < x_t: issues.append(f"X={x}(需≥{x_t})")
            if ct < ct_t: issues.append(f"CT={ct}(需≥{ct_t})")
        if mr != mr_t: issues.append(f"MR={mr}(需{mr_t})")
        if issues:
            print(f"  ❌ {DAYS[day_idx]}: {', '.join(issues)}")
            ok = False
        else:
            print(f"  ✓ {DAYS[day_idx]}: X={x}, CT={ct}, MR={mr}")
    return ok


# ============================================================
# Main
# ============================================================
def main():
    filepath = '/apdcephfs_tj5/share_302216743/poplarzhang/paiban_yl/6.1-6.7排班.xlsx'

    print("=" * 50)
    print("医院医生排班程序 v2")
    print("=" * 50)

    df, doctors = load_all(filepath)
    counts = count_existing(df)

    print(f"\n加载 {len(doctors)} 位医生")

    global _doctors_ref
    _doctors_ref = doctors

    print("\n[Phase 1] 固定排班...")
    phase1(doctors, counts)
    print("[Phase 2] 追加排班(不含MR)...")
    phase2(doctors, counts)
    print("[Phase 3] 有配额医生...")
    phase3(doctors, counts)
    print("[Phase 4] 灵活配额医生...")
    phase4(doctors, counts)

    print("\n分配后:")
    for d in range(5):
        c = counts[d]
        print(f"  {DAYS[d]}: X={c['X']}, CT={c['CT']}, MR={c['MR']}")

    ok = validate(counts)

    print("\n[Phase 4.5] 优化连续班种...")
    iters = phase6_5_optimize(doctors, counts)
    print(f"  优化{iters}轮")

    print("\n[Phase 4.6] 补 MR 短缺(两人三班 swap)...")
    fix_iters = fix_mr_shortfall(doctors, counts)
    print(f"  尝试{fix_iters}轮")
    print("  分配后:")
    for d in range(5):
        c = counts[d]
        print(f"    {DAYS[d]}: X={c['X']}, CT={c['CT']}, MR={c['MR']}")
    ok = validate(counts)

    print("\n[Phase 5] 第一阶段辅助...")
    s1 = phase5(doctors, counts)
    print(f"  分配 {len(s1)} 个")

    print("[Phase 6] 第二阶段辅助...")
    s2 = phase6(doctors, s1)
    print(f"  分配 {len(s2)} 个")

    print("\n[Phase 7] 输出...")
    phase7(doctors, df, s1, s2, filepath)

    if ok:
        print("\n✅ 排班完成!")
    else:
        print("\n⚠️ 有约束未满足")


if __name__ == '__main__':
    main()
